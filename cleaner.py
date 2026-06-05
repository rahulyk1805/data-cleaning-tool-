# ============================================================
# cleaner.py
# ============================================================
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

# ============================================================
# COMMON KEYWORDS
# ============================================================

NON_PRODUCT_KEYWORDS = [
    'banner',
    'brochure',
    'broucher',
    'sticker',
    'outer bag'
]

# ============================================================
# HELPER FUNCTION
# ============================================================

def blank_row(df, index):
    df.loc[index] = [None] * len(df.columns)

# ============================================================
# FILE TYPE DETECTION
# ============================================================

def detect_file_type(file_path):

    try:
        raw_df = pd.read_excel(file_path, header=None)

        text = ""

        for i in range(min(10, len(raw_df))):
            row = raw_df.iloc[i].astype(str).tolist()
            text += " ".join(row).lower()

        # SALES
        if "material name" in text:
            return "SALES"

        # PURCHASE
        elif "customer name" in text and "uom" in text:
            return "PURCHASE"

        # STOCK
        elif "grand total" in text:
            return "STOCK"

        # JV
        elif "partyname" in text or "narration" in text:
            return "JV"
        # HO EXP
        elif "ho" in os.path.basename(file_path).lower():
            return "HO EXP"
        # SUNDRY CREDITORS
        elif "sundry creditors" in text or "group summary" in text:
            return "SUNDRY CREDITORS"

        else:
            return "UNKNOWN"

    except:
        return "UNKNOWN"

# ============================================================
# SALES CLEANING
# ============================================================

def clean_sales_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)
    cleaned_df = raw_df.copy()

    changes = []

    header_row = None

    for idx, row in raw_df.iterrows():

        vals = [
            str(x).strip().lower()
            for x in row.tolist()
            if pd.notna(x)
        ]

        if 'material name' in vals:
            header_row = idx
            break

    if header_row is None:
        return raw_df, cleaned_df, changes

    headers = raw_df.iloc[header_row].tolist()

    data_df = raw_df.iloc[header_row + 1:].copy()
    data_df.columns = headers

    material_col = None

    for col in data_df.columns:

        if 'material' in str(col).lower():
            material_col = col
            break

    if material_col is None:
        return raw_df, cleaned_df, changes

    for i in data_df.index:

        material = str(data_df.at[i, material_col]).strip()
        lower_mat = material.lower()

        # REMOVE NON PRODUCTS
        if any(k in lower_mat for k in NON_PRODUCT_KEYWORDS):

            blank_row(cleaned_df, i)

            changes.append({
                'row': i + 1,
                'reason': 'Non-product row removed'
            })

            continue

        # TYPO FIXES
        new_material = material

        if 'hariyali df-0.05 kg' in lower_mat:
            new_material = 'Hariyali DF 0.5 kg (20 kg box)'

        elif 'haryali df -500gm' in lower_mat:
            new_material = 'Hariyali DF 0.5 kg (20 kg box)'

        elif 'profit df-30 kg (l) d/m' in lower_mat:
            new_material = 'Profit DF 30 kg drum'

        elif 'profit df (1x30kg) loose drum' in lower_mat:
            new_material = 'Profit DF 30 kg drum'

        if new_material != material:

            cleaned_df.at[
                i,
                cleaned_df.columns.get_loc(material_col)
            ] = new_material

            changes.append({
                'row': i + 1,
                'reason': f'Material corrected: {material} → {new_material}'
            })

    return raw_df, cleaned_df, changes

# ============================================================
# PURCHASE CLEANING
# ============================================================

def clean_purchase_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)
    cleaned_df = raw_df.copy()

    changes = []

    header_row = None

    for idx, row in raw_df.iterrows():

        vals = [
            str(x).strip().lower()
            for x in row.tolist()
            if pd.notna(x)
        ]

        if 'si no.' in vals:
            header_row = idx
            break

    if header_row is None:
        return raw_df, cleaned_df, changes

    headers = raw_df.iloc[header_row].tolist()

    data_df = raw_df.iloc[header_row + 1:].copy()
    data_df.columns = headers

    cols = {str(c).strip().lower(): c for c in data_df.columns}

    material_col = None
    uom_col = None
    qty_col = None
    supplier_col = None

    for c in cols:

        if 'material' in c:
            material_col = cols[c]

        elif c == 'uom':
            uom_col = cols[c]

        elif 'quantity' in c:
            qty_col = cols[c]

        elif 'customer name' in c:
            supplier_col = cols[c]

    for i in data_df.index:

        # TON TO KG
        if uom_col and qty_col:

            uom = str(data_df.at[i, uom_col]).lower()

            if 'ton' in uom:

                try:
                    qty = float(
                        str(data_df.at[i, qty_col]).replace(',', '')
                    )

                    new_qty = qty * 1000

                    cleaned_df.at[
                        i,
                        cleaned_df.columns.get_loc(qty_col)
                    ] = new_qty

                    cleaned_df.at[
                        i,
                        cleaned_df.columns.get_loc(uom_col)
                    ] = 'kg'

                    changes.append({
                        'row': i + 1,
                        'reason': 'Quantity converted from ton to kg'
                    })

                except:
                    pass

        # SUPPLIER TYPO
        if supplier_col:

            supplier = str(data_df.at[i, supplier_col])

            if 'chemichal' in supplier.lower():

                corrected = supplier.replace(
                    'CHEMICHAL',
                    'CHEMICAL'
                )

                corrected = corrected.replace(
                    'Chemichal',
                    'Chemical'
                )

                cleaned_df.at[
                    i,
                    cleaned_df.columns.get_loc(supplier_col)
                ] = corrected

                changes.append({
                    'row': i + 1,
                    'reason': 'Supplier spelling corrected'
                })

    return raw_df, cleaned_df, changes

# ============================================================
# STOCK CLEANING
# ============================================================

def clean_stock_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)
    cleaned_df = raw_df.copy()

    changes = []

    for idx, row in raw_df.iterrows():

        material = str(row[0]).strip().lower() \
            if pd.notna(row[0]) else ''

        qty = row[2] if len(row) > 2 else None
        value = row[4] if len(row) > 4 else None

        # REMOVE NON PRODUCTS
        if any(k in material for k in NON_PRODUCT_KEYWORDS):

            blank_row(cleaned_df, idx)

            changes.append({
                'row': idx + 1,
                'reason': 'Non-product row removed'
            })

            continue

        qty_blank = pd.isna(qty) or str(qty).strip() == ''
        value_present = pd.notna(value) and str(value).strip() != ''

        # REMOVE SUBTOTALS
        if qty_blank and value_present:

            if 'grand total' not in material:

                blank_row(cleaned_df, idx)

                changes.append({
                    'row': idx + 1,
                    'reason': 'Subtotal/category removed'
                })

    return raw_df, cleaned_df, changes

# ============================================================
# JV CLEANING
# ============================================================

def clean_jv_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)
    cleaned_df = raw_df.copy()

    changes = []

    header_row = None

    for idx, row in raw_df.iterrows():

        vals = [
            str(x).strip().lower()
            for x in row.tolist()
            if pd.notna(x)
        ]

        if 'si no.' in vals:
            header_row = idx
            break

    if header_row is None:
        return raw_df, cleaned_df, changes

    headers = raw_df.iloc[header_row].tolist()

    data_df = raw_df.iloc[header_row + 1:].copy()
    data_df.columns = headers

    cols = {str(c).strip().lower(): c for c in data_df.columns}

    party_col = None
    particulars_col = None
    narration_col = None
    credit_col = None

    for c in cols:

        if 'partyname' in c:
            party_col = cols[c]

        elif 'particulars' in c:
            particulars_col = cols[c]

        elif 'narration' in c:
            narration_col = cols[c]

        elif 'credit' in c:
            credit_col = cols[c]

    exclusion_keywords = [
        'lodging',
        'telephone',
        'postage',
        'stationery',
        'printing',
        'cartage',
        'freight',
        'transport',
        'motor car',
        'vehicle repair',
        'tyres',
        'tds',
        'cgst',
        'sgst',
        'igst',
        'gst',
        'round off',
        'cash discount',
        'advertising',
        'hotel',
        'repair charges'
    ]

    narration_exclusions = [
        'party merged',
        'balance transfer'
    ]

    for i in data_df.index:

        party = str(data_df.at[i, party_col]).lower() \
            if party_col else ''

        particulars = str(data_df.at[i, particulars_col]).lower() \
            if particulars_col else ''

        narration = str(data_df.at[i, narration_col]).lower() \
            if narration_col else ''

        credit_amt = 0

        if credit_col:

            try:
                credit_raw = str(
                    data_df.at[i, credit_col]
                ).replace(',', '')

                credit_amt = float(credit_raw)

            except:
                credit_amt = 0

        include_row = (
            ('jaishil' in party) and
            (('mumbai' in party) or ('bombay' in party)) and
            ('jaishil' not in particulars) and
            (credit_amt > 0)
        )

        exclude_row = False

        if any(k in particulars for k in exclusion_keywords):
            exclude_row = True

        if any(k in narration for k in narration_exclusions):
            exclude_row = True

        if not include_row or exclude_row:

            blank_row(cleaned_df, i)

            changes.append({
                'row': i + 1,
                'reason': 'JV row excluded'
            })

    return raw_df, cleaned_df, changes
# ============================================================
# HO EXP CLEANING
# ============================================================

def clean_ho_exp_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)

    cleaned_df = raw_df.copy()

    changes = []

    total_row = None

    for i in range(len(raw_df)):

        row_text = " ".join(
            [
                str(x).lower()
                for x in raw_df.iloc[i].tolist()
                if pd.notna(x)
            ]
        )

        if "grand total" in row_text:

            total_row = i
            break

    if total_row is None:

        changes.append({
            "row": 0,
            "reason": "Grand Total row not found"
        })

        return raw_df, cleaned_df, changes

    numeric_columns = []

    for col in range(len(raw_df.columns)):

        count_numeric = 0

        for row in range(total_row):

            try:

                value = str(
                    raw_df.iat[row, col]
                ).replace(",", "")

                float(value)

                count_numeric += 1

            except:
                pass

        if count_numeric >= 3:

            numeric_columns.append(col)

    for col in numeric_columns:

        calculated_total = 0

        for row in range(total_row):

            try:

                value = str(
                    raw_df.iat[row, col]
                ).replace(",", "")

                calculated_total += float(value)

            except:
                pass

        try:

            grand_total_value = str(
                raw_df.iat[total_row, col]
            ).replace(",", "")

            grand_total_value = float(grand_total_value)

        except:

            continue

        if round(calculated_total, 2) != round(grand_total_value, 2):

            changes.append({
                "row": total_row + 1,
                "reason":
                f"Column {col + 1} total mismatch "
                f"(Calculated={calculated_total:.2f}, "
                f"Grand Total={grand_total_value:.2f})"
            })

    return raw_df, cleaned_df, changes
# ============================================================
# SUNDRY CREDITORS CLEANING
# ============================================================

def clean_sundry_creditors_file(file_path):

    raw_df = pd.read_excel(file_path, header=None)
    cleaned_df = raw_df.copy()

    changes = []

    company_name = "jaishil sulphur & chemical industries"

    for i in range(len(cleaned_df)):

        for j in range(len(cleaned_df.columns)):

            value = str(cleaned_df.iat[i, j]).strip()

            if value.lower() == company_name:

                cleaned_df.iat[i, j] = ""

                changes.append({
                    'row': i + 1,
                    'reason': 'Company name removed from creditors column'
                })

    return raw_df, cleaned_df, changes

# ============================================================
# SAVE WORKBOOK
# ============================================================

def save_cleaned_workbook(
    raw_df,
    cleaned_df,
    changes,
    output_path,
    cleaned_sheet_name
):

    wb = Workbook()

    # RAW SHEET
    ws1 = wb.active
    ws1.title = 'Original_Raw'

    for r in raw_df.itertuples(index=False):
        ws1.append(list(r))

    # CLEANED SHEET
    ws2 = wb.create_sheet(cleaned_sheet_name)

    for r in cleaned_df.itertuples(index=False):
        ws2.append(list(r))

    # CHANGES LOG
    ws3 = wb.create_sheet('Changes_Log')

    headers = ['Row Number', 'Reason']

    for col, h in enumerate(headers, start=1):

        cell = ws3.cell(row=1, column=col)
        cell.value = h
        cell.font = Font(bold=True)

    for idx, change in enumerate(changes, start=2):

        ws3.cell(
            row=idx,
            column=1,
            value=change['row']
        )

        ws3.cell(
            row=idx,
            column=2,
            value=change['reason']
        )

    wb.save(output_path)

# ============================================================
# SUMMARY TXT
# ============================================================

def generate_summary_txt(
    filename,
    file_type,
    changes,
    output_path
):

    with open(output_path, "w", encoding="utf-8") as f:

        f.write("DATA CLEANING SUMMARY\n")
        f.write("=" * 50 + "\n\n")

        f.write(f"File Name : {filename}\n")
        f.write(f"Detected Type : {file_type}\n")
        f.write(f"Total Changes : {len(changes)}\n\n")

        f.write("CHANGES MADE:\n")
        f.write("-" * 50 + "\n")

        if len(changes) == 0:

            f.write("No changes were required.\n")

        else:

            for idx, change in enumerate(changes, start=1):

                f.write(
                    f"{idx}. Row {change['row']} → "
                    f"{change['reason']}\n"
                )


